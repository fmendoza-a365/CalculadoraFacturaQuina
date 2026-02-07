import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class QuinaBillingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QuinaBilling App - Automatización de Facturación")
        self.root.geometry("600x500")
        self.root.configure(bg="#f0f2f5")

        # Variables
        self.folder_path = tk.StringVar()
        
        # UI Setup
        self.create_widgets()

    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self.root, bg="#0078d4", height=60)
        header_frame.pack(fill="x")
        title_lbl = tk.Label(header_frame, text="QuinaBilling Automatización", bg="#0078d4", fg="white", font=("Segoe UI", 16, "bold"))
        title_lbl.pack(pady=15)

        # Main Content
        main_frame = tk.Frame(self.root, bg="#f0f2f5", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Instrucciones
        info_lbl = tk.Label(main_frame, text="Selecciona la carpeta mensual que contiene los archivos RDC y DDC.", bg="#f0f2f5", font=("Segoe UI", 10))
        info_lbl.pack(anchor="w", pady=(0, 10))

        # Folder Selection
        folder_frame = tk.Frame(main_frame, bg="white", padx=10, pady=10, relief="flat")
        folder_frame.pack(fill="x")
        
        tk.Label(folder_frame, text="Carpeta del Mes:", bg="white", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        
        entry_frame = tk.Frame(folder_frame, bg="white")
        entry_frame.pack(fill="x", pady=5)
        
        self.entry_path = tk.Entry(entry_frame, textvariable=self.folder_path, font=("Segoe UI", 10), state="readonly")
        self.entry_path.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        btn_browse = tk.Button(entry_frame, text="Examinar...", command=self.browse_folder, bg="#e1e1e1", font=("Segoe UI", 9))
        btn_browse.pack(side="right")

        # Process Button
        self.btn_process = tk.Button(main_frame, text="GENERAR FACTURA", command=self.run_process, bg="#107c10", fg="white", font=("Segoe UI", 12, "bold"), height=2, state="disabled")
        self.btn_process.pack(fill="x", pady=20)

        # Log Area
        lb_log = tk.Label(main_frame, text="Registro de Actividad:", bg="#f0f2f5", font=("Segoe UI", 9, "bold"))
        lb_log.pack(anchor="w")
        
        self.log_text = tk.Text(main_frame, height=10, font=("Consolas", 9), state="disabled", bg="white", relief="flat")
        self.log_text.pack(fill="both", expand=True)

        # Status Bar
        self.status_var = tk.StringVar()
        self.status_var.set("Listo")
        status_bar = tk.Label(self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W, bg="#e1e1e1", font=("Segoe UI", 8))
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.root.update()

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)
            self.btn_process.config(state="normal", bg="#107c10")
            self.log(f"Carpeta seleccionada: {folder_selected}")

    def run_process(self):
        folder = self.folder_path.get()
        if not folder:
            messagebox.showwarning("Advertencia", "Por favor selecciona una carpeta.")
            return

        self.btn_process.config(state="disabled", text="PROCESANDO...")
        try:
            self.process_billing(folder)
            messagebox.showinfo("Éxito", "¡Factura Generada Correctamente!\nVerifica el archivo 'FACTURA_FINAL.xlsx' en la carpeta.")
        except Exception as e:
            self.log(f"❌ ERROR CRÍTICO: {str(e)}")
            messagebox.showerror("Error", f"Ocurrió un error:\n{str(e)}")
        finally:
            self.btn_process.config(state="normal", text="GENERAR FACTURA")

    def process_billing(self, folder):
        self.log("🚀 Iniciando procesamiento...")
        
        # 1. Encontrar archivos
        rdc_files = glob.glob(os.path.join(folder, "RDC_*.xlsx"))
        ddc_files = glob.glob(os.path.join(folder, "DDC_*.xlsx"))

        if not rdc_files: raise Exception("No se encontró archivo RDC_*.xlsx")
        self.log(f"Archivo RDC: {os.path.basename(rdc_files[0])}")
        if not ddc_files: raise Exception("No se encontraron archivos DDC_*.xlsx")
        self.log(f"Archivos DDC: {len(ddc_files)}")

        # 2. Procesar RDC (24h Window)
        self.log("⏳ Calculando Regla 24 Horas (RDC)...")
        df_rdc = pd.read_excel(rdc_files[0], usecols=["ID", "F.Inicio Chat", "ID Chat"])
        df_rdc.dropna(subset=["ID", "F.Inicio Chat"], inplace=True)
        df_rdc["F.Inicio Chat"] = pd.to_datetime(df_rdc["F.Inicio Chat"])
        df_rdc.sort_values(by=["ID", "F.Inicio Chat"], inplace=True)
        
        # Vectorized Lag Calculation
        df_rdc["Prev_ID"] = df_rdc["ID"].shift(1)
        df_rdc["Prev_Time"] = df_rdc["F.Inicio Chat"].shift(1)
        
        # Condiciones: Nuevo ID O Diferencia > 24h
        time_diff = (df_rdc["F.Inicio Chat"] - df_rdc["Prev_Time"]).dt.total_seconds() / 3600.0
        
        # Es cobrable si (Es primer ID) O (Diferencia >= 24)
        # Nota: Primera fila siempre es True para 'Prev_ID != ID'
        is_new_id = df_rdc["ID"] != df_rdc["Prev_ID"]
        is_new_window = time_diff >= 24.0
        
        df_rdc["Es_Cobrable"] = (is_new_id | is_new_window).astype(int)
        
        total_q_hsm = df_rdc["Es_Cobrable"].sum()
        self.log(f"✅ Q HSM (Conversaciones) Calculado: {total_q_hsm:,.0f}")

        # 3. Procesar DDC (Mensajes)
        self.log("⏳ Procesando Mensajes y Reglas de Negocio (DDC)...")
        
        # Carga optimizada iterativa
        dfs = []
        for f in ddc_files:
            dfs.append(pd.read_excel(f, usecols=["ID Chat", "Mensaje", "Fecha Hora", "Tipo"]))
        df_ddc = pd.concat(dfs, ignore_index=True)
        
        df_ddc["Fecha Hora"] = pd.to_datetime(df_ddc["Fecha Hora"])
        # Forzar ID Chat a string para join correcto
        df_ddc["ID Chat"] = df_ddc["ID Chat"].astype(str)
        # Convertir a string para búsquedas de texto
        df_ddc["Mensaje"] = df_ddc["Mensaje"].astype(str).str.lower()
        df_ddc["Tipo"] = df_ddc["Tipo"].astype(str).str.upper().str.strip()

        # Encontrar Hitos por Chat
        self.log("... Analizando Agentes y Créditos ...")
        
        # GroupBy para encontrar MIN(Fecha) de notificación
        agente_times = df_ddc[df_ddc["Tipo"] == "NOTIFICATION"].groupby("ID Chat")["Fecha Hora"].min()
        
        # GroupBy para encontrar MIN(Fecha) de trigger credito
        # Trigger: "podrás evaluar si tienes un crédito"
        credito_mask = df_ddc["Mensaje"].str.contains("podrás evaluar si tienes un crédito", na=False)
        credito_times = df_ddc[credito_mask].groupby("ID Chat")["Fecha Hora"].min()

        # Mapear hitos de vuelta al dataframe principal
        df_ddc["Time_Agente"] = df_ddc["ID Chat"].map(agente_times)
        df_ddc["Time_Credito"] = df_ddc["ID Chat"].map(credito_times)

        # Lógica de Facturabilidad (Vectorizada)
        # Mensaje es Facturable SI:
        # 1. (Time < Time_Agente) O (Time_Agente es NA)
        # 2. Y (Time < Time_Credito) O (Time_Credito es NA)
        
        cond_antes_agente = df_ddc["Time_Agente"].isna() | (df_ddc["Fecha Hora"] < df_ddc["Time_Agente"])
        cond_antes_credito = df_ddc["Time_Credito"].isna() | (df_ddc["Fecha Hora"] < df_ddc["Time_Credito"])
        
        df_ddc["Es_Facturable"] = (cond_antes_agente & cond_antes_credito).astype(int)
        
        total_q_mensajes = df_ddc["Es_Facturable"].sum()
        self.log(f"✅ Q Mensajes (Facturables) Calculado: {total_q_mensajes:,.0f}")

        # 4. Generar Reporte Excel
        self.generate_excel_report(folder, total_q_hsm, total_q_mensajes)

    def generate_excel_report(self, folder, q_hsm, q_mensajes):
        self.log("📊 Generando archivo 'FACTURA_FINAL.xlsx'...")
        path = os.path.join(folder, "FACTURA_FINAL.xlsx")
        
        # Tarifas (Hardcoded por ahora según imagen, podrían ser inputs)
        FEE_MENSUAL = 760.00
        TARIFA_HSM = 0.0770
        TARIFA_MSG = 0.0228
        
        total_hsm_money = q_hsm * TARIFA_HSM
        total_msg_money = q_mensajes * TARIFA_MSG
        subtotal = FEE_MENSUAL + total_hsm_money + total_msg_money
        igv = subtotal * 0.18
        total_facturar = subtotal + igv

        # Crear Excel con Openppyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Factura"

        # Estilos
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Rojo
        sub_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarillo
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Estructura Datos
        data = [
            ["CONCEPTOS", "ABR / CANTIDAD", "MONTO S/", "OBSERVACIONES"], # Header
            ["Fee Mensual", 1, FEE_MENSUAL, "Fee Mensual Broker Whatsapp API Oficial"],
            ["Q HSM", q_hsm, "Calculado", "Conversaciones 24h (Regla Service)"],
            ["Tarifa por HSM", TARIFA_HSM, "Tarifa", "Según adenda N° 2"],
            ["TOTAL HSM", "", total_hsm_money, ""],
            ["Q Mensajes", q_mensajes, "Calculado", "Total Bot - Agente - Conversación Crédito"],
            ["Tarifa por mensajes", TARIFA_MSG, "Tarifa", ""],
            ["TOTAL MENSAJES", "", total_msg_money, ""],
            ["SUB TOTAL", "", subtotal, ""],
            ["IGV (18%)", "", igv, ""],
            ["TOTAL A FACTURAR", "", total_facturar, ""]
        ]

        # Escribir filas
        for i, row in enumerate(data, start=2):
            for j, val in enumerate(row, start=2): # Start col B (2)
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = border
                
                # Formatos moneda
                if isinstance(val, (int, float)) and i > 2 and j == 4: # Columna Monto
                     cell.number_format = '"S/ " #,##0.00'
                
                # Headers y Colores
                if i == 2: # Header Row
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Filas amarillas "Totales"
                if row[0] in ["Fee Mensual", "TOTAL HSM", "TOTAL MENSAJES"]:
                     if j in [2,3,4]: cell.fill = sub_fill

                # Negritas finales
                if row[0] in ["SUB TOTAL", "TOTAL A FACTURAR"]:
                     cell.font = bold_font

        # Ajuste de ancho
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50

        wb.save(path)
        self.log("✅ Archivo guardado con éxito.")

if __name__ == "__main__":
    root = tk.Tk()
    app = QuinaBillingApp(root)
    root.mainloop()
