import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class QuinaCalculator:
    """
    Clase principal para la lógica de facturación.
    Encapsula las reglas de negocio de Quina para procesar archivos RDC y DDC,
    aplicando ventanas de 24h, lógica de crédito y tarifas escalonadas.
    """
    def __init__(self):
        # Constantes de Configuración
        self.FEE_MENSUAL = 760.00
        self.TARIFA_HSM = 0.077
        self.META_FREE_TIER = 1000
        
        # Estado Interno
        self.df_rdc = None
        self.df_ddc = None
        self.df_detalle = None
        
        # Métricas de Facturación
        self.hsm_bruto = 0
        self.hsm_credito = 0
        self.total_q_hsm = 0
        
        self.mensajes_bruto = 0
        self.mensajes_agente = 0
        self.mensajes_credito = 0
        self.total_q_mensajes = 0
        self.total_facturar = 0

    def process_data(self, rdc_source, ddc_sources):
        """
        Función principal de procesamiento.
        rdc_source: ruta al archivo Excel o DataFrame de RDC
        ddc_sources: lista de rutas a archivos Excel o lista de DataFrames de DDC
        """
        self._process_rdc(rdc_source)
        self._process_ddc(ddc_sources)
        return self.get_summary()

    def _process_rdc(self, source):
        # Carga de Datos
        if isinstance(source, pd.DataFrame):
            df = source.copy()
        else:
            df = pd.read_excel(source, usecols=["ID", "F.Inicio Chat", "ID Chat", "Tipificación Chat"])

        # Preprocesamiento
        df.dropna(subset=["ID", "F.Inicio Chat"], inplace=True)
        df["F.Inicio Chat"] = pd.to_datetime(df["F.Inicio Chat"])
        df.sort_values(by=["ID", "F.Inicio Chat"], inplace=True)
        
        # Lógica de Ventana de 24h
        df["Prev_ID"] = df["ID"].shift(1)
        df["Prev_Time"] = df["F.Inicio Chat"].shift(1)
        time_diff = (df["F.Inicio Chat"] - df["Prev_Time"]).dt.total_seconds() / 3600.0
        
        is_new_id = df["ID"] != df["Prev_ID"]
        is_new_window = time_diff >= 24.0
        
        df["Es_Cobrable"] = (is_new_id | is_new_window).astype(int)
        
        # Detección de Crédito (Basado en Tipificación)
        df["ID Chat"] = df["ID Chat"].astype(str)
        mask_tipif_credito = df["Tipificación Chat"].astype(str).str.contains("evalú", case=False, na=False)
        chats_con_credito_tipif = set(df[mask_tipif_credito]["ID Chat"].unique())
        df["Es_Credito"] = df["ID Chat"].isin(chats_con_credito_tipif)

        self.df_rdc = df

        # Cálculos Iniciales de HSM
        self.hsm_bruto = df["Es_Cobrable"].sum()
        # Solo contar crédito si era originalmente cobrable
        self.hsm_credito = df[(df["Es_Cobrable"] == 1) & (df["Es_Credito"])].shape[0]
        # Total Inicial (El paso DDC refina esto, pero se mantiene consistente con la lógica original)

    def _process_ddc(self, sources):
        # Carga de Datos
        dfs = []
        if isinstance(sources, list):
            for source in sources:
                if isinstance(source, pd.DataFrame):
                    dfs.append(source)
                else:
                    dfs.append(pd.read_excel(source, usecols=["ID Chat", "Mensaje", "Fecha Hora", "Tipo"]))
        elif isinstance(sources, pd.DataFrame): # Manejar DataFrame único
             dfs.append(sources)
        
        # Si no hay DDC, manejar ordenadamente
        if not dfs:
            self.total_q_hsm = max(0, self.hsm_bruto - self.hsm_credito - self.META_FREE_TIER)
            self._prepare_simple_detail()
            return

        df_ddc = pd.concat(dfs, ignore_index=True)
        df_ddc["Fecha Hora"] = pd.to_datetime(df_ddc["Fecha Hora"])
        df_ddc["ID Chat"] = df_ddc["ID Chat"].astype(str)
        df_ddc["Tipo"] = df_ddc["Tipo"].astype(str).str.upper().str.strip()
        df_ddc["Mensaje"] = df_ddc["Mensaje"].astype(str).str.lower()

        # Identificar Marcas de Tiempo de Agente y Crédito
        agente_times = df_ddc[df_ddc["Tipo"] == "NOTIFICATION"].groupby("ID Chat")["Fecha Hora"].min()

        credito_mask = (
            df_ddc["Mensaje"].str.contains("evalúa si tienes un crédito", na=False) |
            df_ddc["Mensaje"].str.contains("evalua si tienes un credito", na=False) |
            df_ddc["Mensaje"].str.contains("3. evalúa", na=False) |
            df_ddc["Mensaje"].str.contains("3. evalua", na=False)
        )
        credito_times = df_ddc[credito_mask].groupby("ID Chat")["Fecha Hora"].min()

        df_ddc["Time_Agente"] = df_ddc["ID Chat"].map(agente_times)
        df_ddc["Time_Credito"] = df_ddc["ID Chat"].map(credito_times)

        cond_antes_agente = df_ddc["Time_Agente"].isna() | (df_ddc["Fecha Hora"] < df_ddc["Time_Agente"])
        cond_antes_credito = df_ddc["Time_Credito"].isna() | (df_ddc["Fecha Hora"] < df_ddc["Time_Credito"])

        df_ddc["Es_Facturable"] = (cond_antes_agente & cond_antes_credito).astype(int)
        
        self.df_ddc = df_ddc

        # Calcular Métricas
        self.total_q_mensajes = df_ddc["Es_Facturable"].sum()
        self.mensajes_bruto = len(df_ddc)
        
        cond_post_agente = ~cond_antes_agente
        self.mensajes_agente = cond_post_agente.sum()

        cond_post_credito = cond_antes_agente & (~cond_antes_credito)
        self.mensajes_credito = cond_post_credito.sum()

        # Cálculo Final de HSM
        # Igual que el original: hsm_bruto - hsm_credito - 1000
        self.total_q_hsm = max(0, self.hsm_bruto - self.hsm_credito - self.META_FREE_TIER)

        self._prepare_detailed_report(cond_antes_agente, cond_antes_credito)

    def _prepare_simple_detail(self):
        """Usado cuando no se proporcionan archivos DDC"""
        df = self.df_rdc[["ID Chat", "F.Inicio Chat", "Tipificación Chat", "Es_Cobrable"]].copy()
        df["Fecha_Dia"] = df["F.Inicio Chat"].dt.date
        df["Es_Credito"] = self.df_rdc["Es_Credito"].astype(int)
        # Completar columnas faltantes con 0 o NaT
        for col in ["Mensajes_Bruto", "Mensajes_Post_Agente", "Mensajes_Post_Credito", "Mensajes_Facturables"]:
            df[col] = 0
        df["Time_Agente"] = pd.NaT
        df["Time_Credito"] = pd.NaT
        self.df_detalle = df

    def _prepare_detailed_report(self, cond_antes_agente, cond_antes_credito):
        # 1. Totales Facturables: Conteo de mensajes que cumplen todas las condiciones de cobro
        ddc_counts = self.df_ddc.groupby("ID Chat")["Es_Facturable"].sum().reset_index()
        ddc_counts.rename(columns={"Es_Facturable": "Mensajes_Facturables"}, inplace=True)
        
        # 2. Mensajes Bruto por Chat: Total absoluto de mensajes registrados en el periodo
        ddc_bruto = self.df_ddc.groupby("ID Chat").size().reset_index(name="Mensajes_Bruto")
        
        # 3. Mensajes Post-Agente: Mensajes descartados por ocurrir después de la derivación a humano
        self.df_ddc["Es_Post_Agente"] = (~cond_antes_agente).astype(int)
        ddc_post_agente = self.df_ddc.groupby("ID Chat")["Es_Post_Agente"].sum().reset_index()
        ddc_post_agente.rename(columns={"Es_Post_Agente": "Mensajes_Post_Agente"}, inplace=True)
        
        # 4. Mensajes Post-Crédito: Mensajes descartados por ocurrir después de la notificación de crédito
        self.df_ddc["Es_Post_Credito"] = (cond_antes_agente & (~cond_antes_credito)).astype(int)
        ddc_post_credito = self.df_ddc.groupby("ID Chat")["Es_Post_Credito"].sum().reset_index()
        ddc_post_credito.rename(columns={"Es_Post_Credito": "Mensajes_Post_Credito"}, inplace=True)
        
        # 5. Metadatos Temporales: Marcas de tiempo de primera derivación a agente o crédito
        ddc_meta = self.df_ddc[["ID Chat", "Time_Agente", "Time_Credito"]].groupby("ID Chat").first().reset_index()
        
        # Unificación de Métricas (Merge): Combinar todos los cálculos en un solo DataFrame de vista DDC
        ddc_view = pd.merge(ddc_counts, ddc_bruto, on="ID Chat", how="left")
        ddc_view = pd.merge(ddc_view, ddc_post_agente, on="ID Chat", how="left")
        ddc_view = pd.merge(ddc_view, ddc_post_credito, on="ID Chat", how="left")
        ddc_view = pd.merge(ddc_view, ddc_meta, on="ID Chat", how="left")
        ddc_view["ID Chat"] = ddc_view["ID Chat"].astype(str)
        
        # Preparación de Vista RDC: Seleccionar columnas relevantes del Resumen Diario
        rdc_view = self.df_rdc[["ID Chat", "F.Inicio Chat", "Tipificación Chat", "Es_Cobrable", "Es_Credito"]].copy()
        rdc_view["ID Chat"] = rdc_view["ID Chat"].astype(str)

        # Fusión Maestra: Unir información de RDC (base) con métricas detalladas de DDC
        df_detalle = pd.merge(rdc_view, ddc_view, on="ID Chat", how="left")
        
        # Limpieza de Datos: Rellenar valores nulos resultantes del merge con ceros (para enteros)
        cols_to_fill = ["Mensajes_Facturables", "Mensajes_Bruto", "Mensajes_Post_Agente", "Mensajes_Post_Credito"]
        df_detalle[cols_to_fill] = df_detalle[cols_to_fill].fillna(0).astype(int)
        
        df_detalle["Es_Credito"] = df_detalle["Es_Credito"].astype(int)
        df_detalle["Fecha_Dia"] = df_detalle["F.Inicio Chat"].dt.date
        
        # Selección de Columnas Finales: Filtrar y ordenar las columnas para el reporte de auditoría
        self.df_detalle = df_detalle[[
            "ID Chat", "Fecha_Dia", "F.Inicio Chat", "Tipificación Chat",
            "Es_Cobrable", "Es_Credito", "Mensajes_Bruto",
            "Mensajes_Post_Agente", "Mensajes_Post_Credito", "Mensajes_Facturables",
            "Time_Agente", "Time_Credito"
        ]]

    def get_summary(self):
        return {
            "Total HSM Final": self.total_q_hsm,
            "Total Mensajes Final": self.total_q_mensajes,
            "HSM Bruto": self.hsm_bruto,
            "HSM Credito": self.hsm_credito,
            "Mensajes Bruto": self.mensajes_bruto,
            "Mensajes Agente": self.mensajes_agente,
            "Mensajes Credito": self.mensajes_credito
        }

    def generate_excel_report(self):
        """Genera los bytes del archivo Excel"""
        output = io.BytesIO()
        
        # Lógica de Cálculo para Facturación
        def calcular_costo_mensajes(cantidad):
            # Cálculo de tarifa escalonada según el volumen total de mensajes
            if cantidad <= 0: return 0.0
            elif cantidad <= 9999: return cantidad * 0.0456
            elif cantidad <= 99999: return cantidad * 0.0380
            elif cantidad <= 249999: return cantidad * 0.0304
            else: return cantidad * 0.0228

        total_hsm_money = self.total_q_hsm * self.TARIFA_HSM
        total_msg_money = calcular_costo_mensajes(self.total_q_mensajes)
        subtotal = self.FEE_MENSUAL + total_hsm_money + total_msg_money
        igv = subtotal * 0.18
        total_facturar = subtotal + igv
        
        # Generación del Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Factura"
        
        # Estilos
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        sub_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Determinar tarifa aplicable para mostrar
        q_mensajes = self.total_q_mensajes
        if q_mensajes <= 9999: tarifa_msg_aplicable = 0.0456
        elif q_mensajes <= 99999: tarifa_msg_aplicable = 0.0380
        elif q_mensajes <= 249999: tarifa_msg_aplicable = 0.0304
        else: tarifa_msg_aplicable = 0.0228

        data = [
            ["CONCEPTOS", "ABR / CANTIDAD", "MONTO S/", "OBSERVACIONES"],
            ["Fee Mensual", 1, self.FEE_MENSUAL, "Fee Mensual Broker Whatsapp API Oficial"],
            ["", "", "", ""],
            ["CÁLCULO HSM (Detallado)", "", "", ""],
            ["HSM Bruto (Total Conversaciones 24h)", self.hsm_bruto, "", "Antes de descuentos"],
            ["(-) HSM Opción 3 (Evalúa tu Crédito)", -self.hsm_credito, "", "Sesiones que derivaron a crédito"],
            ["(-) HSM Meta Free Tier", -self.META_FREE_TIER, "", "1,000 conversaciones gratuitas Meta"],
            ["Q HSM Neto Facturable", self.total_q_hsm, "Calculado", "HSM a cobrar después de descuentos"],
            ["Tarifa por HSM", self.TARIFA_HSM, "Tarifa", "Según adenda N° 2"],
            ["TOTAL HSM", "", total_hsm_money, ""],
            ["", "", "", ""],
            ["CÁLCULO MENSAJES (Detallado)", "", "", ""],
            ["Mensajes Bruto (Total)", self.mensajes_bruto, "", "Todos los mensajes del periodo"],
            ["(-) Mensajes Post-Agente", -self.mensajes_agente, "", "Mensajes después de pase a humano"],
            ["(-) Mensajes Post-Crédito", -self.mensajes_credito, "", "Mensajes después de trigger crédito"],
            ["Q Mensajes Neto Facturable", self.total_q_mensajes, "Calculado", "Mensajes a cobrar después de descuentos"],
            ["Tarifa por mensajes", tarifa_msg_aplicable, "Tarifa", "Tarifa escalonada aplicada al volumen"],
            ["TOTAL MENSAJES", "", total_msg_money, ""],
            ["", "", "", ""],
            ["SUB TOTAL", "", subtotal, ""],
            ["IGV (18%)", "", igv, ""],
            ["TOTAL A FACTURAR", "", total_facturar, ""]
        ]

        for i, row in enumerate(data, start=1):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = border
                if isinstance(val, (int, float)) and i > 1 and j == 3: cell.number_format = '"S/ " #,##0.00'
                if i == 1:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')
                if row[0] in ["Fee Mensual", "TOTAL HSM", "TOTAL MENSAJES"] and j in [1,2,3]:
                    cell.fill = sub_fill
                if row[0] in ["SUB TOTAL", "TOTAL A FACTURAR"] or row[0] in ["CÁLCULO HSM (Detallado)", "CÁLCULO MENSAJES (Detallado)"]:
                    cell.font = bold_font

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50

        # Hoja 2: Detalle de Auditoría - Generación de tabla con desglose detallado por chat
        if self.df_detalle is not None and not self.df_detalle.empty:
            ws2 = wb.create_sheet("Detalle Auditoría")
            headers = [
                "ID Chat", "Fecha (Día)", "F.Inicio (RDC)", "Tipificación Chat",
                "Es HSM Bruto? (1=Sí)", "Tuvo Crédito? (1=Sí)",
                "Mensajes Bruto", "(-) Mensajes Post-Agente", "(-) Mensajes Post-Crédito",
                "Mensajes Facturables (Neto)", "Fecha Corte Agente", "Fecha Corte Crédito"
            ]
            
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal='center')
            
            rows = self.df_detalle.values.tolist()
            for i, row_data in enumerate(rows, start=2):
                for j, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=i, column=j, value=val)
                    if j == 1: cell.number_format = '@'

            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 12
            ws2.column_dimensions['C'].width = 20
            ws2.column_dimensions['D'].width = 30
            ws2.column_dimensions['E'].width = 18
            ws2.column_dimensions['F'].width = 18
            ws2.column_dimensions['G'].width = 15
            ws2.column_dimensions['H'].width = 22
            ws2.column_dimensions['I'].width = 22
            ws2.column_dimensions['J'].width = 20
            ws2.column_dimensions['K'].width = 20
            ws2.column_dimensions['L'].width = 20

        wb.save(output)
        return output.getvalue()
